import oracledb
import logging
from datetime import date, timedelta, datetime
from collections import OrderedDict
from django.conf import settings
from django.shortcuts import render, redirect
from django.views.generic import TemplateView
from django.views import View
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
import pandas as pd
import statistics
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from apps.accounts.decorators import role_required
from .forms import UploadFileForm
from .models import ParametrosRiesgo, ParametrosRiesgoSarL

logger = logging.getLogger(__name__)

# Definición explícita del orden de los indicadores, según lo solicitado.
INDICADORES_LIMITES_ORDER = [
    'Total Cartera',
    'Maximo Cartera Consumo',
    'Maximo Cartera Microcredito',
    'Maximo Cartera Comercial',
    'Máximo de Cartera Otras Garantías',
    'Máximo de Cartera Sin Garantías',
    'Máximo de Cartera Hipoteca',
    'Máximo de Cartera Rentas',
    'Máximo de Cartera Otras Admisibles',
    'Máximo de cartera Fondos de garantía',
    'Maximo Indicador Morosidad',
    'Limite de Acumulacion',
    'Concentración por deudor',
    'Concentración por línea de crédito',
    'Concentración por clasificación de cartera',
    'Concentración por plazo al vencimiento',
    'Concentración por zona',
    'Concentración por sector económico',
]

INDICADORES_GRAFICABLES_ORDER = [
    'Indicador De Cartera Vencida Por Temporalidad (Mora)',
    'Indicador De Cartera Vencida Por Riesgo (Riesgo)',
    'Indicador De Cartera Improductiva',
    'Rodamiento General De Cartera',
    'Rodamiento De Cartera A-B',
    'Indicador De Fallo',
    'Reverso General del Deterioro',
    'Indicador De Cobertura',
    'Relacion Riesgo / Mora',
    'Cartera Castigada',
    'Cartera Reestructurada',
    'Calidad De Cosechas',
    'Crecimiento Cartera Bruta',
]

# Mapa de código a nombre, usado para el procesamiento de datos.
INDICADOR_MAP = {
    '1': 'Total Cartera',
    '2': 'Maximo Cartera Consumo',
    '3': 'Maximo Cartera Microcredito',
    '4': 'Maximo Cartera Comercial',
    '5': 'Máximo de Cartera Otras Garantías',
    '6': 'Máximo de Cartera Sin Garantías',
    '7': 'Máximo de Cartera Hipoteca',
    '8': 'Máximo de Cartera Rentas',
    '9': 'Máximo de Cartera Otras Admisibles',
    '10': 'Máximo de cartera Fondos de garantía',
    '11': 'Maximo Indicador Morosidad',
    '12': 'Limite de Acumulacion',
    '13': 'Concentración por deudor',
    '14': 'Concentración por línea de crédito',
    '15': 'Concentración por clasificación de cartera',
    '16': 'Concentración por plazo al vencimiento',
    '17': 'Concentración por zona',
    '18': 'Concentración por sector económico',
    '1G': 'Indicador De Cartera Vencida Por Temporalidad (Mora)',
    '2G': 'Indicador De Cartera Vencida Por Riesgo (Riesgo)',
    '3G': 'Indicador De Cartera Improductiva',
    '4G': 'Rodamiento General De Cartera',
    '5G': 'Rodamiento De Cartera A-B',
    '6G': 'Indicador De Fallo',
    '7G': 'Reverso General del Deterioro',
    '8G': 'Indicador De Cobertura',
    '9G': 'Relacion Riesgo / Mora',
    '10G': 'Cartera Castigada',
    '11G': 'Cartera Reestructurada',
    '12G': 'Calidad De Cosechas',
    '13G': 'Crecimiento Cartera Bruta',
}

def get_fecha_corte():
    """
    Calcula la fecha de corte según las reglas del procedimiento
    """
    hoy = date.today()
    primer_dia_mes_actual = hoy.replace(day=1)
    corte = primer_dia_mes_actual - timedelta(days=1)
    return corte.strftime('%Y%m')

def obtener_datos_sarc_sp(fecha_corte):
    try:
        db = settings.DATABASES['oracle']
        dsn = f"{db['HOST']}:{db['PORT']}/{db['NAME']}"
        with oracledb.connect(user=db['USER'], password=db['PASSWORD'], dsn=dsn) as conn:
            with conn.cursor() as cursor:
                ref_cursor_out = cursor.var(oracledb.CURSOR)
                cursor.callproc('SP_INDICADORESRIESGOS', [fecha_corte, ref_cursor_out])
                cur = ref_cursor_out.getvalue()
                if cur:
                    cols = [c[0] for c in cur.description]
                    return [dict(zip(cols, row)) for row in cur]
        return []
    except Exception as e:
        logger.error(f"Error en obtener_datos_sarc_sp: {e}", exc_info=True)
        return []

def parse_fecha_mes(fecha_valor):
    if not isinstance(fecha_valor, str) or '-' not in fecha_valor:
        return None, None
    try:
        mes_abbr, anio = fecha_valor.strip().split('-')
        mes_num_map = {
            'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
            'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12
        }
        mes_num = mes_num_map.get(mes_abbr.upper().strip())
        if mes_num and anio.strip().isdigit():
            anio_str = anio.strip()
            anio_completo = int(f"20{anio_str}") if len(anio_str) == 2 else int(anio_str)
            fecha_obj = date(anio_completo, mes_num, 1)
            month_key = f"{mes_abbr.upper()}-{anio_str[-2:]}"
            return fecha_obj, month_key
    except (ValueError, KeyError, AttributeError) as e:
        logger.warning(f"Error parseando fecha '{fecha_valor}': {e}")
    return None, None

@method_decorator(login_required, name='dispatch')
@method_decorator(role_required(allowed_roles=['Riesgos']), name='dispatch')
class DashboardSarcView(TemplateView):
    template_name = "risk/dashboard/dashboardSarc.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fecha_corte = get_fecha_corte()
        raw_data = obtener_datos_sarc_sp(fecha_corte)
        parametros_manuales = {p.indicador_codigo: p for p in ParametrosRiesgo.objects.all()}
        
        unique_months = OrderedDict()
        for row in raw_data:
            _, month_key = parse_fecha_mes(row.get('MES'))
            if month_key and month_key not in unique_months:
                unique_months[month_key] = None
        
        months_columns = [{'key': k, 'display': k} for k in unique_months.keys()]
        latest_month_key = months_columns[-1]['key'] if months_columns else None

        limites_pivot, indicadores_pivot = self._pivot_data(raw_data, parametros_manuales, latest_month_key)

        context['limites_table_data'] = self._format_pivot_for_template(limites_pivot, months_columns, INDICADORES_LIMITES_ORDER)
        context['indicadores_table_data'] = self._format_pivot_for_template(indicadores_pivot, months_columns, INDICADORES_GRAFICABLES_ORDER)
        
        context['tabla_riesgo_limites'] = self._generar_tabla_riesgo(raw_data, parametros_manuales, False, latest_month_key)
        context['tabla_riesgo_indicadores'] = self._generar_tabla_riesgo(raw_data, parametros_manuales, True, latest_month_key)

        context['months_columns'] = months_columns
        return context

    def _pivot_data(self, raw_data, parametros_manuales, latest_month_key):
        pivot_data = {}
        for row in raw_data:
            indicador_code = str(row.get('INDICADOR', '')).strip()
            nombre_descriptivo = INDICADOR_MAP.get(indicador_code)
            _, month_key = parse_fecha_mes(row.get('MES'))
            if not (nombre_descriptivo and month_key):
                continue

            parametro = parametros_manuales.get(indicador_code)
            valor_raw = row.get('VALOR')
            valor = None
            if valor_raw is not None:
                try:
                    valor = float(str(valor_raw).replace(',', '.'))
                except (ValueError, TypeError):
                    valor = None

            if parametro and parametro.valor_override is not None and parametro.valor_override_mes:
                if month_key == parametro.valor_override_mes.upper():
                    valor = parametro.valor_override
            
            if valor is None:
                continue

            pivot_data.setdefault(nombre_descriptivo, {})[month_key] = valor
        
        limites_pivot = {k: v for k, v in pivot_data.items() if k in INDICADORES_LIMITES_ORDER}
        indicadores_pivot = {k: v for k, v in pivot_data.items() if k in INDICADORES_GRAFICABLES_ORDER}
        return limites_pivot, indicadores_pivot

    def _generar_tabla_riesgo(self, raw_data, parametros_manuales, is_graphable, latest_month_key):
        resultados = []
        map_code_to_name = {k: v for k, v in INDICADOR_MAP.items()}
        map_name_to_code = {v: k for k, v in map_code_to_name.items()}
        order_list = INDICADORES_GRAFICABLES_ORDER if is_graphable else INDICADORES_LIMITES_ORDER

        for nombre_indicador in order_list:
            ind_code = map_name_to_code.get(nombre_indicador)
            if not ind_code:
                continue

            parametro = parametros_manuales.get(ind_code)

            datos_indicador_historicos = []
            for d in raw_data:
                if INDICADOR_MAP.get(str(d.get('INDICADOR', '')).strip()) != nombre_indicador or d.get('MES') is None:
                    continue
                fecha_obj, month_key = parse_fecha_mes(d.get('MES'))
                if not fecha_obj or not month_key:
                    continue

                valor = d.get('VALOR')
                if valor is not None:
                    try:
                        valor = float(str(valor).replace(',', '.'))
                    except (ValueError, TypeError):
                        valor = None

                if parametro and parametro.valor_override is not None and parametro.valor_override_mes:
                    if month_key == parametro.valor_override_mes.upper():
                        valor = parametro.valor_override

                if valor is not None:
                    d_copy = dict(d)
                    d_copy['parsed_date'] = fecha_obj
                    d_copy['valor_override_applied'] = True if (parametro and parametro.valor_override is not None and parametro.valor_override_mes and month_key == parametro.valor_override_mes.upper()) else False
                    d_copy['VALOR'] = valor
                    d_copy['parsed_month_key'] = month_key
                    datos_indicador_historicos.append(d_copy)

            valor_actual = 0
            if datos_indicador_historicos:
                datos_indicador_historicos.sort(key=lambda x: x['parsed_date'], reverse=True)
                registro_mas_reciente = datos_indicador_historicos[0]
                valor_actual = registro_mas_reciente.get('VALOR', 0) or 0

                month_key = registro_mas_reciente.get('parsed_month_key')
                if parametro and parametro.valor_override is not None and parametro.valor_override_mes:
                    if month_key == parametro.valor_override_mes.upper():
                        valor_actual = registro_mas_reciente.get('VALOR', valor_actual)

            elif parametro and parametro.valor_override is not None and parametro.valor_override_mes:
                if parametro.valor_override_mes.upper() == latest_month_key:
                    valor_actual = parametro.valor_override

            apetito = parametro.apetito if parametro else None
            tolerancia = parametro.tolerancia if parametro else None
            capacidad = parametro.capacidad if parametro else None

            riesgo = 'N/A'
            if parametro and all(p is not None for p in [parametro.apetito, parametro.tolerancia, parametro.capacidad]):
                apetito, tolerancia, capacidad = parametro.apetito, parametro.tolerancia, parametro.capacidad
                if valor_actual < apetito:
                    riesgo = 'Bajo'
                elif valor_actual < capacidad and valor_actual > apetito:
                    riesgo = 'Medio'
                else:
                    riesgo = 'Alto'

            resultados.append({
                "INDICADOR": nombre_indicador,
                "RIESGO": riesgo,
                "APETITO": apetito,
                "TOLERANCIA": tolerancia,
                "CAPACIDAD": capacidad,
            })
        return resultados

    def _format_pivot_for_template(self, pivot_data, months_columns, order_list):
        table_data = []
        for nombre in order_list:
            row_data = {'nombre': nombre, 'monthly_values': {}}
            for mc in months_columns:
                row_data['monthly_values'][mc['key']] = pivot_data.get(nombre, {}).get(mc['key'], '-')
            table_data.append(row_data)
        return table_data

@method_decorator(login_required, name='dispatch')
@method_decorator(role_required(allowed_roles=['Riesgos']), name='dispatch')
class UploadParametrosRiesgoView(View):
    template_name = 'risk/upload_parametros.html'
    form_class = UploadFileForm
    
    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "Formulario inválido.")
            return render(request, self.template_name, {'form': form})
        
        file = request.FILES['file']
        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file, dtype={'indicador_codigo': str})
            elif file.name.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file, dtype={'indicador_codigo': str})
            else:
                messages.error(request, "Formato de archivo no soportado. Use CSV o Excel.")
                return redirect(request.path)

            df.columns = [col.strip().lower() for col in df.columns]
            updated_count, created_count = 0, 0

            def _parse_decimal(value):
                if pd.isna(value):
                    return None
                if isinstance(value, str):
                    value = value.strip().replace(',', '.')
                    if not value:
                        return None
                try:
                    return float(value)
                except (TypeError, ValueError):
                    return None

            def _normalize_month(value):
                if pd.isna(value):
                    return None
                if isinstance(value, (pd.Timestamp, datetime, date)):
                    eng_to_esp = {
                        'JAN': 'ENE', 'FEB': 'FEB', 'MAR': 'MAR', 'APR': 'ABR',
                        'MAY': 'MAY', 'JUN': 'JUN', 'JUL': 'JUL', 'AUG': 'AGO',
                        'SEP': 'SEP', 'OCT': 'OCT', 'NOV': 'NOV', 'DEC': 'DIC'
                    }
                    eng_abbr = value.strftime('%b').upper()
                    esp_abbr = eng_to_esp.get(eng_abbr, eng_abbr)
                    return f"{esp_abbr}-{value.strftime('%y')}"

                raw_text = str(value).strip()
                if not raw_text:
                    return None

                _, month_key = parse_fecha_mes(raw_text)
                if month_key:
                    return month_key

                try:
                    parsed_dt = pd.to_datetime(raw_text, dayfirst=True, errors='coerce')
                except Exception:
                    parsed_dt = None

                if pd.isna(parsed_dt):
                    return None
                return _normalize_month(parsed_dt)

            for _, row in df.iterrows():
                indicador_codigo = row.get('indicador_codigo')
                if pd.isna(indicador_codigo):
                    continue
                
                indicador_codigo = str(indicador_codigo).strip()
                if not indicador_codigo:
                    continue

                sistema_val = row.get('sistema', 'SARC')
                if pd.isna(sistema_val):
                    sistema_val = 'SARC'
                sistema = str(sistema_val).strip().upper()
                if sistema not in ['SARC', 'SARL']:
                    sistema = 'SARC'
                model_cls = ParametrosRiesgo if sistema == 'SARC' else ParametrosRiesgoSarL

                defaults = {}
                allowed_cols = ['apetito', 'tolerancia', 'capacidad', 'valor_override', 'valor_override_mes']
                if hasattr(model_cls, 'orden'):
                    allowed_cols.append('orden')

                for col in allowed_cols:
                    if col in row and pd.notna(row[col]):
                        if col == 'valor_override_mes':
                            month_key = _normalize_month(row[col])
                            if month_key:
                                defaults[col] = month_key
                        elif col == 'orden':
                            orden_val = str(row[col]).strip().upper()
                            if orden_val.startswith('DESC'):
                                orden_val = 'DESC'
                            elif orden_val.startswith('ASC'):
                                orden_val = 'ASC'
                            defaults[col] = orden_val
                        elif col == 'valor_override':
                            val = _parse_decimal(row[col])
                            if val is not None:
                                defaults[col] = val
                        else:
                            val = _parse_decimal(row[col])
                            defaults[col] = val if val is not None else row[col]

                if not defaults:
                    continue

                _, created = model_cls.objects.update_or_create(
                    indicador_codigo=indicador_codigo,
                    defaults=defaults
                )
                if created: created_count += 1
                else: updated_count += 1
            
            messages.success(request, f"Archivo procesado. {created_count} registros creados, {updated_count} actualizados.")
        except Exception as e:
            logger.error(f"Error al procesar el archivo de parámetros: {e}", exc_info=True)
            messages.error(request, f"Ocurrió un error al procesar el archivo: {e}")
        
        return redirect(request.path)

class DashboardSarcDataJsonView(TemplateView):
    def get(self, request, *args, **kwargs):
        fecha_corte = get_fecha_corte()
        raw_data = obtener_datos_sarc_sp(fecha_corte)
        parametros_manuales = {p.indicador_codigo: p for p in ParametrosRiesgo.objects.all()}
        map_name_to_code = {v: k for k, v in INDICADOR_MAP.items()}

        indicators_data = {}
        unique_dates = {}
        
        for row in raw_data:
            indicador_code = str(row.get('INDICADOR', '')).strip()
            if 'G' not in indicador_code: continue

            nombre_descriptivo = INDICADOR_MAP.get(indicador_code)
            valor = row.get('VALOR')
            fecha_valor = row.get('MES')

            if not all([nombre_descriptivo, valor is not None, fecha_valor]): continue
            
            fecha_obj, month_key = parse_fecha_mes(fecha_valor)
            if not fecha_obj or not month_key: continue
            
            try:
                valor_float = float(str(valor).replace(',', '.'))
            except (ValueError, TypeError):
                valor_float = 0
            
            unique_dates[month_key] = fecha_obj
            indicators_data.setdefault(nombre_descriptivo, {})[month_key] = {
                'fecha': fecha_obj, 'valor': valor_float
            }
        
        sorted_dates = sorted(unique_dates.items(), key=lambda x: x[1])
        chart_x_labels = [k for k, _ in sorted_dates]
        
        trend_data = []
        for nombre in INDICADORES_GRAFICABLES_ORDER:
            monthly_data = indicators_data.get(nombre, {})
            points = [{'fecha': label, 'valor': monthly_data.get(label, {}).get('valor')} for label in chart_x_labels]
            
            indicador_code = map_name_to_code.get(nombre)
            parametro = parametros_manuales.get(indicador_code)
            
            apetito, capacidad = None, None
            if parametro:
                apetito = parametro.apetito
                capacidad = parametro.capacidad

            trend_data.append({
                "nombre_limite": nombre, 
                "data_points": points,
                "apetito": apetito,
                "capacidad": capacidad
            })
        
        return JsonResponse({"trend_data": trend_data, "chart_labels": chart_x_labels}, safe=False)

def descargar_plantilla_parametros(request):
    """
    Genera y sirve un archivo Excel que sirve como plantilla para la carga de parámetros de riesgo.
    El archivo contiene solo los encabezados de las columnas esperadas.
    """
    headers = [
        'sistema',  # Opcional: SARC (por defecto) o SARL
        'indicador_codigo',
        'apetito',
        'tolerancia',
        'capacidad',
        'valor_override',
        'valor_override_mes',
        'orden'  # Solo para SARL (ASC o DESC)
    ]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Parámetros"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_align = Alignment(horizontal='center')

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = len(header_title) + 5

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_parametros_riesgo.xlsx"'
    
    wb.save(response)
    
    return response

#? -----------------------------------------------------------------------------------------------------------------
#? VISTA PARA SARL
#? -----------------------------------------------------------------------------------------------------------------

def obtener_datos_sarl_sp(fecha_corte):
    try:
        db = settings.DATABASES['oracle']
        dsn = f"{db['HOST']}:{db['PORT']}/{db['NAME']}"
        with oracledb.connect(user=db['USER'], password=db['PASSWORD'], dsn=dsn) as conn:
            with conn.cursor() as cursor:
                ref_cursor_out = cursor.var(oracledb.CURSOR)
                cursor.callproc('SP_INDICADORESRIESGOSLI', [fecha_corte, ref_cursor_out])
                cur = ref_cursor_out.getvalue()
                if cur:
                    cols = [c[0] for c in cur.description]
                    return [dict(zip(cols, row)) for row in cur]
        return []
    except Exception as e:
        logger.error(f"Error en obtener_datos_sarl_sp: {e}", exc_info=True)
        return []

#* Definición de los indicadores para SARL
SARL_INDICADORES_NO_GRAFICABLES_ORDER = [
    'Liquidez sobre los depósitos',
    'Cubrimiento primario fondo de liquidez',
    'Relación entre las obligaciones financieras y el pasivo total',
    'Relación entre activos liquidos ampliados a depósitos de corto plazo',
    'Relación entre la brecha de liquidez y los activos liquidos netos',
    'Relacion de Activos Netos Liquidos',
    'Porcentaje de cupos disponibles de crédito con la banca comercial',
    'Tasa de renovación CDAT',
    'IRL Interno',
]

SARL_INDICADORES_GRAFICABLES_ORDER = [
    'Estructura de balance',
    'Activo Productivo / Activo',
    'Financiamiento de activos con pasivos a corto plazo',
    'Financiamiento de activos con pasivos de una sola entidad financiera',
    'Financiamiento de activos con pasivos de entidades financieras',
    'VaR Avis',
    'VaR.aportes',
    'Retiro máx. probable cuenta ahorros',
    'Indicador de mora ahorro contractual',
    'Concentración de cartera (20 principales deudores)',
    'Concentración de depositos (20 principales ahorradores)',
    'Concentración Avis',
    'Concentración Ahorro contractual',
    'Concentración CDAT según dias al vencimiento',
    'Concentración CDAT por asociado',
    'Concentración CDAT por tasa',
    'Concentración mayor depósitante persona natural o jurídica',
]

# ATENCIÓN: Los códigos de indicador ('SL1', 'SG1', etc.) son placeholders.
# Deben ser verificados y ajustados según los códigos reales devueltos por el SP_INDICADORESRIESGOSLI.
SARL_INDICADOR_MAP = {
    '1': 'Liquidez sobre los depósitos',
    '2': 'Cubrimiento primario fondo de liquidez',
    '3': 'Relación entre las obligaciones financieras y el pasivo total',
    '4': 'Relación entre activos liquidos ampliados a depósitos de corto plazo',
    '5': 'Relación entre la brecha de liquidez y los activos liquidos netos',
    '6': 'Relacion de Activos Netos Liquidos',
    '7': 'Porcentaje de cupos disponibles de crédito con la banca comercial',
    '8': 'Tasa de renovación CDAT',
    '9': 'IRL Interno',
    '1G': 'Estructura de balance',
    '2G': 'Activo Productivo / Activo',
    '3G': 'Financiamiento de activos con pasivos a corto plazo',
    '4G': 'Financiamiento de activos con pasivos de una sola entidad financiera',
    '5G': 'Financiamiento de activos con pasivos de entidades financieras',
    '6G': 'VaR Avis',
    '7G': 'VaR.aportes',
    '8G': 'Retiro máx. probable cuenta ahorros',
    '9G': 'Indicador de mora ahorro contractual',
    '10G': 'Concentración de cartera (20 principales deudores)',
    '11G': 'Concentración de depositos (20 principales ahorradores)',
    '12G': 'Concentración Avis',
    '13G': 'Concentración Ahorro contractual',
    '14G': 'Concentración CDAT según dias al vencimiento',
    '15G': 'Concentración CDAT por asociado',
    '16G': 'Concentración CDAT por tasa',
    '17G': 'Concentración mayor depósitante persona natural o jurídica',
}


@method_decorator(login_required, name='dispatch')
@method_decorator(role_required(allowed_roles=['Riesgos']), name='dispatch')
class DashboardSarLView(TemplateView):
    template_name = "risk/dashboard/dashboardSarL.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fecha_corte = get_fecha_corte()
        raw_data = obtener_datos_sarl_sp(fecha_corte)
        parametros_manuales = {p.indicador_codigo: p for p in ParametrosRiesgoSarL.objects.all()}
        
        unique_months = OrderedDict()
        for row in raw_data:
            _, month_key = parse_fecha_mes(row.get('MES'))
            if month_key and month_key not in unique_months:
                unique_months[month_key] = None
        
        months_columns = [{'key': k, 'display': k} for k in unique_months.keys()]
        latest_month_key = months_columns[-1]['key'] if months_columns else None

        limites_pivot, indicadores_pivot = self._pivot_data(raw_data, parametros_manuales, latest_month_key)

        context['limites_table_data'] = self._format_pivot_for_template(limites_pivot, months_columns, SARL_INDICADORES_NO_GRAFICABLES_ORDER)
        context['indicadores_table_data'] = self._format_pivot_for_template(indicadores_pivot, months_columns, SARL_INDICADORES_GRAFICABLES_ORDER)
        
        context['tabla_riesgo_limites'] = self._generar_tabla_riesgo(raw_data, parametros_manuales, False, latest_month_key)
        context['tabla_riesgo_indicadores'] = self._generar_tabla_riesgo(raw_data, parametros_manuales, True, latest_month_key)

        context['months_columns'] = months_columns
        return context

    def _pivot_data(self, raw_data, parametros_manuales, latest_month_key):
        pivot_data = {}
        for row in raw_data:
            indicador_code = str(row.get('INDICADOR', '')).strip()
            nombre_descriptivo = SARL_INDICADOR_MAP.get(indicador_code)
            _, month_key = parse_fecha_mes(row.get('MES'))
            if not (nombre_descriptivo and month_key):
                continue

            parametro = parametros_manuales.get(indicador_code)
            valor_raw = row.get('VALOR')
            valor = None
            if valor_raw is not None:
                try:
                    valor = float(str(valor_raw).replace(',', '.'))
                except (ValueError, TypeError):
                    valor = None

            if parametro and parametro.valor_override is not None and parametro.valor_override_mes:
                if month_key == parametro.valor_override_mes.upper():
                    valor = parametro.valor_override
            
            if valor is None:
                continue

            pivot_data.setdefault(nombre_descriptivo, {})[month_key] = valor
        
        limites_pivot = {k: v for k, v in pivot_data.items() if k in SARL_INDICADORES_NO_GRAFICABLES_ORDER}
        indicadores_pivot = {k: v for k, v in pivot_data.items() if k in SARL_INDICADORES_GRAFICABLES_ORDER}
        return limites_pivot, indicadores_pivot

    def _generar_tabla_riesgo(self, raw_data, parametros_manuales, is_graphable, latest_month_key):
        resultados = []
        map_name_to_code = {v: k for k, v in SARL_INDICADOR_MAP.items()}
        order_list = SARL_INDICADORES_GRAFICABLES_ORDER if is_graphable else SARL_INDICADORES_NO_GRAFICABLES_ORDER

        for nombre_indicador in order_list:
            ind_code = map_name_to_code.get(nombre_indicador)
            if not ind_code:
                continue

            parametro = parametros_manuales.get(ind_code)
            
            datos_indicador_historicos = []
            for d in raw_data:
                if SARL_INDICADOR_MAP.get(str(d.get('INDICADOR', '')).strip()) != nombre_indicador or d.get('MES') is None:
                    continue
                fecha_obj, month_key = parse_fecha_mes(d.get('MES'))
                if not fecha_obj or not month_key:
                    continue

                valor = d.get('VALOR')
                if valor is not None:
                    try:
                        valor = float(str(valor).replace(',', '.'))
                    except (ValueError, TypeError):
                        valor = None

                if parametro and parametro.valor_override is not None and parametro.valor_override_mes:
                    if month_key == parametro.valor_override_mes.upper():
                        valor = parametro.valor_override

                if valor is not None:
                    d_copy = dict(d)
                    d_copy['parsed_date'] = fecha_obj
                    d_copy['VALOR'] = valor
                    d_copy['parsed_month_key'] = month_key
                    datos_indicador_historicos.append(d_copy)

            valor_actual = 0
            if datos_indicador_historicos:
                datos_indicador_historicos.sort(key=lambda x: x['parsed_date'], reverse=True)
                registro_mas_reciente = datos_indicador_historicos[0]
                valor_actual = registro_mas_reciente.get('VALOR', 0) or 0

                month_key = registro_mas_reciente.get('parsed_month_key')
                if parametro and parametro.valor_override is not None and parametro.valor_override_mes:
                    if month_key == parametro.valor_override_mes.upper():
                        valor_actual = registro_mas_reciente.get('VALOR', valor_actual)
            
            elif parametro and parametro.valor_override is not None and parametro.valor_override_mes:
                if parametro.valor_override_mes.upper() == latest_month_key:
                    valor_actual = parametro.valor_override

            apetito = parametro.apetito if parametro else None
            tolerancia = parametro.tolerancia if parametro else None
            capacidad = parametro.capacidad if parametro else None

            riesgo = 'N/A'
            if parametro and apetito is not None and tolerancia is not None:
                r0 = abs(tolerancia - apetito) / 4
                orden_raw = (parametro.orden or '').strip().upper()
                is_desc = orden_raw.startswith('DESC')
                # Descendente: parte desde apetito; Ascendente: parte desde tolerancia
                r1 = apetito if is_desc else tolerancia
                thresholds = [r1, r1 + r0, r1 + 2 * r0, r1 + 3 * r0, r1 + 4 * r0]
                labels = ['MINIMO', 'BAJO', 'MEDIO', 'ALTO', 'MUY ALTO'] if is_desc else ['MUY ALTO', 'ALTO', 'MEDIO', 'BAJO', 'MINIMO']
                riesgo = labels[0]
                for t, label in zip(thresholds, labels):
                    if valor_actual >= t:
                        riesgo = label
            
            resultados.append({
                "INDICADOR": nombre_indicador,
                "RIESGO": riesgo,
                "APETITO": apetito,
                "TOLERANCIA": tolerancia,
                "CAPACIDAD": capacidad,
            })
        return resultados

    def _format_pivot_for_template(self, pivot_data, months_columns, order_list):
        table_data = []
        for nombre in order_list:
            row_data = {'nombre': nombre, 'monthly_values': {}}
            for mc in months_columns:
                row_data['monthly_values'][mc['key']] = pivot_data.get(nombre, {}).get(mc['key'], '-')
            table_data.append(row_data)
        return table_data

class DashboardSarLDataJsonView(TemplateView):
    def get(self, request, *args, **kwargs):
        fecha_corte = get_fecha_corte()
        raw_data = obtener_datos_sarl_sp(fecha_corte)
        parametros_manuales = {p.indicador_codigo: p for p in ParametrosRiesgoSarL.objects.all()}
        map_name_to_code = {v: k for k, v in SARL_INDICADOR_MAP.items()}

        indicators_data = {}
        unique_dates = {}
        
        for row in raw_data:
            indicador_code = str(row.get('INDICADOR', '')).strip()
            nombre_descriptivo = SARL_INDICADOR_MAP.get(indicador_code)
            
            if not nombre_descriptivo or nombre_descriptivo not in SARL_INDICADORES_GRAFICABLES_ORDER:
                continue

            valor = row.get('VALOR')
            fecha_valor = row.get('MES')

            if not all([valor is not None, fecha_valor]): continue
            
            fecha_obj, month_key = parse_fecha_mes(fecha_valor)
            if not fecha_obj or not month_key: continue
            
            try:
                valor_float = float(str(valor).replace(',', '.'))
            except (ValueError, TypeError):
                valor_float = 0
            
            unique_dates[month_key] = fecha_obj
            indicators_data.setdefault(nombre_descriptivo, {})[month_key] = {
                'fecha': fecha_obj, 'valor': valor_float
            }
        
        sorted_dates = sorted(unique_dates.items(), key=lambda x: x[1])
        chart_x_labels = [k for k, _ in sorted_dates]
        
        trend_data = []
        for nombre in SARL_INDICADORES_GRAFICABLES_ORDER:
            monthly_data = indicators_data.get(nombre, {})
            points = [{'fecha': label, 'valor': monthly_data.get(label, {}).get('valor')} for label in chart_x_labels]
            
            indicador_code = map_name_to_code.get(nombre)
            parametro = parametros_manuales.get(indicador_code)
            
            apetito, tolerancia = None, None
            if parametro:
                apetito = parametro.apetito
                tolerancia = parametro.tolerancia

            trend_data.append({
                "nombre_limite": nombre, 
                "data_points": points,
                "apetito": apetito,
                "tolerancia": tolerancia,
            })
        
        return JsonResponse({"trend_data": trend_data, "chart_labels": chart_x_labels}, safe=False)
