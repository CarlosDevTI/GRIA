from datetime import datetime, date, timedelta
import json
from openai import OpenAI
import logging
from django.shortcuts import render
from django.http import HttpResponse
from django.views.decorators.http import require_http_methods
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .oracle_service import *

client = OpenAI(api_key=settings.OPENAI_API_KEY or "")

logger = logging.getLogger(__name__)

#? ------------------------------------
#?    VISTA PRINCIPAL FONDEOS
#? ------------------------------------
def seguimiento_fondeo(request):
    datos = []
    error = None

    if request.method == "POST":
        try: 
            datos = obtener_datos_fondeo()
        except Exception as e:
            logger.error(f"Error al obtener datos de fondeo: {e}", exc_info=True)
            error = "Error al obtener los datos. Por favor, intente nuevamente más tarde."
            datos =[]

    return render(request, "fondeos/fondeos_dashboard.html", {
        "fondeo_list": datos,
        "fecha_generacion": datetime.now(),
        "error": error
    })


#* ---------------------------------------
#*        EXPORTAR A EXCEL FONDEOS
#* ---------------------------------------

def exportar_fondeo_excel(request):
    try:
        datos = obtener_datos_fondeo()

        #* Definir headers
        header_fondeo = [
            'CODSUCURSAL', 'SUCURSAL', 'AHORRAFACIL', 'CONSEMILLA', 'JUNIOR',
            'CONAHORRITO', 'TOTALAHORROS', 'CDAT', 'CONTRACTUAL',
            'TOTCDATCONTRA', 'TOTALCAPTACIONES', 'APORTES',
            'CARTERA', 'CAPAPOCART', 'RECAUDOALCALDIAS'
        ]

        return exportar_a_excel(
            datos,
            f"fondeo_{datetime.now().date()}",
            header_fondeo
        )
    except ValueError as e:
        return HttpResponse(str(e), status=400)
    except Exception as e:
        logger.error(f"Error al exportar fondeo a Excel: {e}", exc_info=True)
        return HttpResponse("Error al generar el reporte.", status=500)
    

#? ----------------------------------------------
#?    VISTA PRINCIPAL ASOCIADOS SIN PRODUCTOS
#? ----------------------------------------------
def asociados_sin_productos_view(request):
    datos = []
    error = None

    if request.method == "POST":
        agencia = request.POST.get('agencia')
        try: 
            datos = asociados_sin_productos(agencia=agencia)
            print(datos)
        except Exception as e:
            logger.error(f"Error al obtener datos de asociados sin productos: {e}", exc_info=True)
            error = "Error al obtener los datos. Por favor, intente nuevamente más tarde."
            datos =[]

    return render(request, "asociados_sin_productos/asociados_sin_productos_dashboard.html", {
        "asociados_list": datos,
        "fecha_generacion": datetime.now(),
        "error": error
    })


#* ------------------------------------------------------
#*    EXPORTAR A EXCEL ASOCIADOS SIN PRODUCTOS
#* ------------------------------------------------------
def exportar_asociados_sin_productos_excel(request):
    if request.method == "POST":
        agencia = request.POST.get('agencia')
        try:
            # Suponiendo que la función puede filtrar por agencia
            datos = asociados_sin_productos(agencia=agencia)

            # Definir headers para el Excel
            header_asociados = ["CORTE", "AGENCIA", "CEDULA", "NOMBRE", "DIRECCION", "TELEFONO", "BARRIO", "MAIL", "CELULAR", "FECHA_VINCULACION", "GESTION", "LINEA"]

            return exportar_a_excel(
                datos,
                f"asociados_sin_producto_{agencia}_{datetime.now().date()}",
                header_asociados
            )
        except ValueError as e:
            return HttpResponse(str(e), status=400)
        except Exception as e:
            logger.error(f"Error al exportar asociados sin productos a Excel: {e}", exc_info=True)
            return HttpResponse("Error al generar el reporte.", status=500)
    
    return HttpResponse("Método no permitido.", status=405)

#? ----------------------------------------------
#?       VISTA DE DETALLES DE APORTES
#? ----------------------------------------------
def detalles_aportes(request):
    context = {
        "fecha_generacion": datetime.now(),
        "aportes_data": None,
        "aportes_data_json": "[]",  # Inicializar como JSON vacío
        "error": None
    }
    meses_orden = ['ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
    context["meses_data"] = [{'nombre': mes} for mes in meses_orden]

    if request.method == "POST":
        try:
            datos_crudos = obtener_datos_aportes()
            if not datos_crudos:
                context['error'] = "No se encontraron datos de aportes para el periodo consultado."
            else:
                # Serializar los datos para que el JS los pueda usar
                context['aportes_data_json'] = json.dumps(datos_crudos)
            
            def procesar_una_sola_fila(fila_datos):
                if not fila_datos:
                    return [{'aportantes': None, 'porcentaje_aportantes': None, 'millones': None, 'porcentaje_millones': None}] * len(meses_orden)
                
                meses_procesados = []
                for mes in meses_orden:
                    meses_procesados.append({
                        'aportantes': fila_datos.get(f'APORTANTES_{mes}'),
                        'porcentaje_aportantes': fila_datos.get(f'PORCEAPORTANTES_{mes}'),
                        'millones': fila_datos.get(f'MILLONES_{mes}'),
                        'porcentaje_millones': fila_datos.get(f'PORCENMILLONES_{mes}'),
                    })
                return meses_procesados

            context["aportes_data"] = datos_crudos
            
            data_keys = [
                "juridico_data", "juridico_barrido_data", "juridico_voluntario_data", "juridico_ninguno_data", "juridico_subtotal_data",
                "mayores_data", "mayores_barrido_data", "mayores_voluntario_data", "mayores_ninguno_data", "mayores_subtotal_data",
                "menores_data", "menores_barrido_data", "menores_voluntario_data", "menores_ninguno_data", "menores_subtotal_data",
                "total_data", "variaciones_data"
            ]

            for i, key in enumerate(data_keys):
                context[key] = procesar_una_sola_fila(datos_crudos[i] if len(datos_crudos) > i else None)

            context["total_registros"] = len(datos_crudos) if datos_crudos else 0

        except Exception as e:
            logger.error(f"Error al obtener datos de aportes: {e}", exc_info=True)
            context['error'] = "Error crítico al procesar los datos de aportes."

    return render(request, "aportes/aportes.html", context)


#* ----------------------------------------------
#*       EXPORTAR A EXCEL DETALLES DE APORTES
#* ----------------------------------------------
def exportar_excel_aportes(request):
    # Obtener datos de aportes usando la función de servicio
    try:
        datos = obtener_datos_aportes()
    except Exception as e:
        logger.error(f"Error al obtener datos de aportes para el excel: {e}", exc_info=True)
        return HttpResponse("Error al obtener los datos para el reporte.", status=500)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle Aportes"

    # Encabezados (ajusta según tus columnas de aportes)
    headers = [
        'Tipo Aporte', 'Abril Aportantes', 'Abril %', 'Abril Millones', 'Abril % Millones',
        'Mayo Aportantes', 'Mayo %', 'Mayo Millones', 'Mayo % Millones',
        'Junio Aportantes', 'Junio %', 'Junio Millones', 'Junio % Millones',
        'Julio Aportantes', 'Julio %', 'Julio Millones', 'Julio % Millones',
        'Agosto Aportantes', 'Agosto %', 'Agosto Millones', 'Agosto % Millones',
        'Septiembre Aportantes', 'Septiembre %', 'Septiembre Millones', 'Septiembre % Millones',
        'Octubre Aportantes', 'Octubre %', 'Octubre Millones', 'Octubre % Millones',
        'Noviembre Aportantes', 'Noviembre %', 'Noviembre Millones', 'Noviembre % Millones',
        'Diciembre Aportantes', 'Diciembre %', 'Diciembre Millones', 'Diciembre % Millones'
    ]

    # Estilos encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_align = Alignment(horizontal='center')

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # Llenar datos
    labels = [
        "JURIDICO", "  BARRIDO", "  VOLUNTARIO", "  NINGUNO", "Aportantes",
        "MAYORES", "  BARRIDO", "  VOLUNTARIO", "  NINGUNO", "Aportantes",
        "MENORES", "  BARRIDO", "  VOLUNTARIO", "  NINGUNO", "Aportantes",
        "ASOCIADOS ACTIVOS", "VARIACIÓN"
    ]
    meses_orden = ['ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
    money_format = '#,##0'
    percent_format = '0.00%'

    for i, row in enumerate(datos, 2):
        # Tipo de registro
        label_index = i - 2
        if label_index < len(labels):
            ws.cell(i, 1, labels[label_index])
        else:
            ws.cell(i, 1, row.get('TIPO') or '')

        col_idx = 2
        for mes in meses_orden:
            ws.cell(i, col_idx, row.get(f'APORTANTES_{mes}') or 0)
            ws.cell(i, col_idx + 1, float(row.get(f'PORCEAPORTANTES_{mes}') or 0) / 100).number_format = percent_format
            ws.cell(i, col_idx + 2, float(row.get(f'MILLONES_{mes}') or 0)).number_format = money_format
            ws.cell(i, col_idx + 3, float(row.get(f'PORCENMILLONES_{mes}') or 0) / 100).number_format = percent_format
            col_idx += 4

    # Ajustar anchos
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = length + 2

    # Respuesta
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="aportes_{datetime.now().date()}.xlsx"'
    wb.save(resp)
    return resp

#* ----------------------------------------------
#*       GENERAR RESUMEN IA DE APORTES
#* ----------------------------------------------
@require_http_methods(["POST"])
def generar_resumen_aportes(request):
    """
    Genera un resumen técnico del reporte de Aportes usando OpenAI (Responses API) y devuelve HTML.
    """
    logger.info("=== INICIANDO generar_resumen_aportes ===")

    if not settings.OPENAI_API_KEY:
        logger.error("OPENAI_API_KEY no está configurada en settings.py")
        return JsonResponse({'error': 'La clave de API de OpenAI no está configurada.'}, status=500)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        logger.exception("JSON inválido en el cuerpo de la petición")
        return JsonResponse({'error': 'JSON inválido.'}, status=400)

    report_data = body.get('data', [])
    if not report_data:
        return JsonResponse({'error': 'No se proporcionaron datos para el resumen.'}, status=400)


    # Opcional: acotar tamaño para evitar prompts gigantes
    report_str = json.dumps(report_data, ensure_ascii=False)
    if len(report_str) > 120_000:  # ~120 KB de texto, ajusta a tus límites
        logger.warning("report_data muy grande; truncando para el prompt")
        report_str = report_str[:120_000] + "\n... (truncado)"

    # Prompts dentro de la FUNCION (system + user)
    system_prompt = (
        "Eres un analista financiero experto. Redacta un concepto técnico "
        "consolidado del reporte de Aportes en formato HTML limpio y semántico."
    )

    user_prompt = f"""
Instrucciones clave:
- Considera únicamente los meses que tienen datos (ignora meses sin información, desde septiembre en adelante).
- Mantén el análisis breve, técnico y conciso.
- Usa etiquetas HTML (<h2>, <p>, <ul>, <li>, <b>) correctamente.
- Resalta con <b> los términos o datos más importantes.
- Evita repetir información irrelevante o redundante.

Estructura esperada:
<h2>Análisis de Tendencias y Puntos Clave</h2>
<ul>
<li>3 hallazgos principales en el comportamiento de los aportes</li>
</ul>

<h2>Recomendaciones</h2>
<ul>
<li>Máximo 3-4 acciones concretas y accionables</li>
</ul>

Datos para el análisis (JSON):
{report_str}
"""

    try:
        # Chat Completions API (modelo sugerido; ajusta si usas otro)
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0.3,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        )

        # Extraer texto de la respuesta
        summary_html = (resp.choices[0].message.content or "").strip()

        if not summary_html:
            logger.error("La respuesta de OpenAI vino vacía")
            return JsonResponse({'error': 'No se obtuvo contenido de la IA.'}, status=502)

        return JsonResponse({'summary': summary_html})

    except Exception as e:
        logger.exception(f"Error al llamar a OpenAI: {e}")
        return JsonResponse({'error': 'Error al generar el resumen con IA.'}, status=500)

#? ==============================
#?  APORTES DETALLE POR CEDULA
#? ==============================

def _create_aportes_detalle_excel(datos):
    """
    Crea un archivo Excel con el detalle de recaudos y lo devuelve como una HttpResponse.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle Aportes por Cédula"

    # --- Encabezados y Estilos ---
    headers = ['Cédula', 'Valor', 'Tipo']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_align = Alignment(horizontal='center')

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Llenar datos ---
    money_format = '#,##0'
    for row_num, row_data in enumerate(datos, 2):
        ws.cell(row=row_num, column=1, value=row_data.get('CEDULA'))
        
        valor_cell = ws.cell(row=row_num, column=2, value=float(row_data.get('VALOR') or 0))
        valor_cell.number_format = money_format
        
        ws.cell(row=row_num, column=3, value=row_data.get('TIPO'))

    # --- Ajustar anchos de columna ---
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # --- Crear y devolver la respuesta ---
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="recaudo_aportes_detalle_{datetime.now().date()}.xlsx"'
    wb.save(resp)
    return resp


def generar_recaudo_aportes_detalle(request):
    """
    Genera un reporte detallado de recaudo de aportes por cédula.

    Procedimiento almacenado: SP_APORTESUNOAUNO
    Parámetros: Fecha Actual, Cursor de salida
    Devuelve: CEDULA, VALOR, TIPO
    """
    if request.method != "POST":
        return HttpResponse("Método no permitido. Use POST para generar el reporte.", status=405)

    try:
        hoy = datetime.now().date()
        f_actual = hoy.strftime('%Y/%m/%d')
        logger.info(f"Generando reporte de recaudo de aportes para la fecha: {f_actual}")

        # Usar la función genérica para ejecutar el procedimiento
        datos = ejecutar_procedimiento('SP_APORTESUNOAUNO', [f_actual])

        if not datos:
            logger.warning("No se encontraron datos de recaudo de aportes para la fecha solicitada.")
            return HttpResponse("No se encontraron datos para generar el reporte.", status=200)

        # Si hay datos, generar y devolver el archivo Excel
        return _create_aportes_detalle_excel(datos)

    except oracledb.DatabaseError as db_err:
        logger.error(f"Error de base de datos al generar el reporte de aportes: {db_err}", exc_info=True)
        return HttpResponse("Error al conectar con la base de datos. Por favor, intente más tarde.", status=500)
    except Exception as e:
        logger.error(f"Error inesperado al generar el reporte de aportes: {e}", exc_info=True)
        return HttpResponse("Ocurrió un error inesperado al generar el reporte.", status=500)