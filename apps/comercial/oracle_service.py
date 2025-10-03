from django.conf import settings
import oracledb
import logging
from datetime import datetime, date, timedelta
import calendar
from django.http import JsonResponse, HttpRequest
from django.views.decorators.http import require_POST


logger = logging.getLogger(__name__)

#? ----------------------------------------------------------------
#?                ABSTRACCION DE LA LÓGICA DE ORACLE
#? -----------------------------------------------------------------
def ejecutar_procedimiento(procedimiento, parametros=[]):
    """
    Función base para ejecutar todos los procedmientos y no repetir código.
    Tengo que validar que todos devuelvan un cursor como último parametro.
    """
    try:
        db= settings.DATABASES['oracle'] # Configurar la base
        dsn = f"{db['HOST']}:{db['PORT']}/{db['NAME']}" # Construir el DSN

        with oracledb.connect(user=db['USER'], password=db['PASSWORD'], dsn=dsn) as conn: # Cerrar automáticamente la conexión.
            with conn.cursor() as cursor:   # Cerrar atumáticamente la conexión.
                ref_cursor_out = cursor.var(oracledb.CURSOR)    # Obtener el resultado del cursor devuelto por el SP.
                parametros_completos = parametros + [ref_cursor_out]    # Agregar el cursor de salida

                cursor.callproc(procedimiento, parametros_completos)    # Ejecutar el procedimiento ('Nombre del procedimiento" , [param1, param2, ..., cursor_out])
                cur = ref_cursor_out.getvalue()  # Obtener el cursor devuelto

                if cur:
                    cols = [c[0] for c in cur.description]
                    return [dict(zip(cols, row)) for row in cur]
                
                return []
            
    except Exception as e:
        logger.error(f"Error en ejecutar_procedimiento {procedimiento}: {e}", exc_info=True)
        raise
#? ----------------------     CONEXION A ORACLE   ------------------------------------

#* --------------------------------------------------------------------------
#*        ABSTRACCION / FUNCION REUTILIZABLE PARA EXPORTAR A EXCEL
#* --------------------------------------------------------------------------
def exportar_a_excel(datos, nombre_archivo, headers=None):
    """
    Exporta datos a Excel de forma genérica.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse

    if not datos:
        raise ValueError("No hay datos para exportar.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    #* Pasarle los encabezados si se proporcionan
    if not headers and datos:
        headers = list(datos[0].keys())

    #* Estilos encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_align = Alignment(horizontal='center')

    #* Escribir encabezados
    for col, header in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=header) #* c -> celda
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
    
    #* Escribir datos
    for row_idx, row_data in enumerate(datos, 2):
        for col_idx, key in enumerate(headers, 1):
            value = row_data.get(key, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    #* Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    #* Crear respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    wb.save(response)

    return response
#* ----------------------     EXPORTAR A EXCEL   ------------------------------------



#? ----------------------------------------------------------------
#?     FUNCIÓN REUTILIZABLE PARA OBETENER LOS DATOS DE FONDEOS
#? ----------------------------------------------------------------
def obtener_datos_fondeo():
    """SP_FONDEO"""
    fecha_generacion = datetime.now()
    hoy = fecha_generacion.date()

    _, ultimo_dia_mes = calendar.monthrange(hoy.year, hoy.month)
    fecha_corte_actual = date(hoy.year, hoy.month, ultimo_dia_mes)
    fecha_corte_anterior = hoy.replace(day=1) - timedelta(days=1)

    # Formato para Oracle
    f_actual = fecha_corte_actual.strftime('%Y/%m/%d')
    print("fecha_actual", f_actual)
    f_generacion = fecha_generacion.strftime('%Y/%m/%d')
    print("fecha_generacion", f_generacion)
    f_anterior = fecha_corte_anterior.strftime('%Y/%m/%d')
    print("fecha_anterior", f_anterior)

    return ejecutar_procedimiento('SP_FONDEO', [f_actual, f_generacion, f_anterior])


#? --------------------------------------------------------------------------------
#?     FUNCION REUTILIZABLE PARA OBTENER LOS DATOS DE ASOCIADOS SIN PRODUCTOS
#? --------------------------------------------------------------------------------
def asociados_sin_productos(agencia):
    """SP_COMEASOSINPRO"""
    return ejecutar_procedimiento('SP_COMEASOSINPRO', [agencia])


#? ----------------------------------------------------------------
#?     FUNCIÓN REUTILIZABLE PARA OBETENER LOS DATOS DE APORTES
#? ----------------------------------------------------------------
def obtener_datos_aportes():
    """SP_DETALLEAPORTES"""
    fecha_generacion = datetime.now()
    hoy = fecha_generacion.date()

    _, ultimo_dia_mes = calendar.monthrange(hoy.year, hoy.month)
    fecha_corte_actual = date(hoy.year, hoy.month, ultimo_dia_mes)
    fecha_corte_anterior = hoy.replace(day=1) - timedelta(days=1)

    # Formato para Oracle
    f_actual = fecha_corte_actual.strftime('%Y/%m/%d')
    print("fecha_actual", f_actual)
    f_generacion = fecha_generacion.strftime('%Y/%m/%d')
    print("fecha_generacion", f_generacion)
    f_anterior = fecha_corte_anterior.strftime('%Y/%m/%d')
    print("fecha_anterior", f_anterior)

    return ejecutar_procedimiento('SP_DETALLEAPORTES', [f_actual, f_generacion, f_anterior])