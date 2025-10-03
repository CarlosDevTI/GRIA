from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import Workbook
import pandas as pd
from django.views.decorators.http import require_POST
from django.contrib import messages
from .models import Indicador, RegistroIndicador, Formula
from .forms import IndicadorForm, RegistroIndicadorForm, FormulaForm

import oracledb

from datetime import datetime, timedelta
from django.db import models

#? -----------------------------------------------
#? FUNCION PARA MOSTRAR LA TABLA DE INDICADORES
#? -----------------------------------------------
def lista_indicadores(request):
    indicadores = Indicador.objects.prefetch_related('formulas__registros').order_by('orden')
    today = datetime.now()
    current_year = today.year
    current_month = today.month
    previous_year = current_year - 1

    meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    headers = []
    for month_num in range(1, 13):
        headers.append({'año': previous_year, 'mes': meses_nombres[month_num-1], 'nombre': meses_nombres[month_num-1]})
    for month_num in range(1, current_month + 1):
        headers.append({'año': current_year, 'mes': meses_nombres[month_num-1], 'nombre': meses_nombres[month_num-1]})

    table_data = []
    for indicador in indicadores:
        indicador_data = {
            'indicador': indicador,
            'formulas': [],
            'rowspan': indicador.formulas.count() or 1
        }
        
        if not indicador.formulas.exists():
            indicador_data['formulas'].append({
                'formula': None,
                'valores': ['-'] * len(headers)
            })
        else:
            for formula in indicador.formulas.order_by('id'):
                registros_map = { (reg.año, reg.mes): reg for reg in formula.registros.all() }
                
                valores = []
                for header in headers:
                    valor_obj = registros_map.get((header['año'], header['mes']))
                    valores.append(valor_obj.valor if valor_obj else '-')
                
                indicador_data['formulas'].append({
                    'formula': formula,
                    'valores': valores
                })
        
        table_data.append(indicador_data)

    context = {
        'table_data': table_data,
        'headers': headers,
        'titulo': 'Indicadores SARLAFT',
    }
    
    return render(request, 'sarlaft/lista_indicadores.html', context)

#? -------------------------------------
def agregar_indicador(request):
    if request.method == 'POST':
        form = IndicadorForm(request.POST)
        if form.is_valid():
            # Solo creamos el indicador
            indicador = form.save()
            messages.success(request, f'Indicador "{indicador.nombre}" creado. Ahora puede agregarle fórmulas.')
            return redirect('agregar_formula', indicador_id=indicador.id)
    else:
        form = IndicadorForm()
    
    return render(request, 'sarlaft/agregar_indicador.html', {
        'form': form,
        'titulo': 'Crear Nuevo Indicador'
    })

#? -------------------------------------
def editar_indicador(request, id):
    indicador = get_object_or_404(Indicador, id=id)
    if request.method == 'POST':
        form = IndicadorForm(request.POST, instance=indicador)
        if form.is_valid():
            form.save()
            messages.success(request, 'Indicador actualizado correctamente.')
            return redirect('lista_indicadores')
    else:
        form = IndicadorForm(instance=indicador)
    
    return render(request, 'sarlaft/editar_indicador.html', {
        'form': form, 
        'indicador': indicador,
        'titulo': 'Editar Indicador'
    })

#? -------------------------------------
def eliminar_indicador(request, id):
    indicador = get_object_or_404(Indicador, id=id)
    
    if request.method == 'POST':
        indicador.delete()
        messages.success(request, 'Indicador y todas sus fórmulas asociadas han sido eliminados.')
        return redirect('lista_indicadores')
    
    return render(request, 'sarlaft/eliminar_indicador.html', {
        'indicador': indicador,
        'titulo': 'Eliminar Indicador'
    })

#? -------------------------------------
def agregar_formula(request, indicador_id):
    indicador = get_object_or_404(Indicador, id=indicador_id)
    
    if request.method == 'POST':
        form = FormulaForm(request.POST)
        if form.is_valid():
            formula = form.save(commit=False)
            formula.indicador = indicador
            formula.save()
            messages.success(request, f'Fórmula agregada al indicador "{indicador.nombre}".')
            # Redirigir a la misma página para ver la nueva fórmula y poder agregar otra
            return redirect('agregar_formula', indicador_id=indicador.id)
    else:
        form = FormulaForm()

    # Obtener las fórmulas existentes para este indicador
    formulas_existentes = Formula.objects.filter(indicador=indicador)

    return render(request, 'sarlaft/agregar_formula.html', {
        'form': form,
        'indicador': indicador,
        'formulas': formulas_existentes,
        'titulo': f'Agregar y ver Fórmulas de {indicador.nombre}'
    })

#? -------------------------------------
def editar_formula(request, formula_id):
    formula = get_object_or_404(Formula, id=formula_id)
    indicador = formula.indicador # Para la navegación y el contexto

    if request.method == 'POST':
        form = FormulaForm(request.POST, instance=formula)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fórmula actualizada correctamente.')
            return redirect('agregar_formula', indicador_id=indicador.id)
    else:
        form = FormulaForm(instance=formula)
    
    return render(request, 'sarlaft/editar_formula.html', {
        'form': form,
        'formula': formula,
        'indicador': indicador,
        'titulo': 'Editar Fórmula'
    })

#? -------------------------------------
def eliminar_formula(request, formula_id):
    formula = get_object_or_404(Formula, id=formula_id)
    indicador_id = formula.indicador.id # Guardar el ID antes de borrar

    if request.method == 'POST':
        formula.delete()
        messages.success(request, 'Fórmula eliminada correctamente.')
        return redirect('agregar_formula', indicador_id=indicador_id)
    
    return render(request, 'sarlaft/eliminar_formula.html', {
        'formula': formula,
        'indicador': formula.indicador,
        'titulo': 'Eliminar Fórmula'
    })

#? -------------------------------------
def gestionar_registros(request, formula_id):
    formula = get_object_or_404(Formula, id=formula_id)
    registros = formula.registros.all()
    
    if request.method == 'POST':
        form = RegistroIndicadorForm(request.POST)
        if form.is_valid():
            año = form.cleaned_data['año']
            mes = form.cleaned_data['mes']
            
            # Update or create logic
            registro, created = RegistroIndicador.objects.update_or_create(
                formula=formula, año=año, mes=mes,
                defaults={'valor': form.cleaned_data['valor']}
            )
            
            if created:
                messages.success(request, 'Registro agregado correctamente.')
            else:
                messages.info(request, 'Registro actualizado correctamente.')
            
            return redirect('gestionar_registros', formula_id=formula_id)
    else:
        form = RegistroIndicadorForm()
    
    return render(request, 'sarlaft/gestionar_registros.html', {
        'formula': formula,
        'registros': registros,
        'form': form,
        'titulo': f'Registros de "{formula.indicador.nombre}"'
    })

#? -------------------------------------
def editar_registro(request, id):
    registro = get_object_or_404(RegistroIndicador, id=id)
    formula_id = registro.formula.id
    if request.method == 'POST':
        form = RegistroIndicadorForm(request.POST, instance=registro)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro actualizado.')
            return redirect('gestionar_registros', formula_id=formula_id)
    else:
        form = RegistroIndicadorForm(instance=registro)
    
    return render(request, 'sarlaft/editar_registro.html', {
        'form': form,
        'registro': registro,
        'indicador': registro.formula.indicador,
        'titulo': 'Editar Registro'
    })

#? -------------------------------------
def eliminar_registro(request, id):
    registro = get_object_or_404(RegistroIndicador, id=id)
    formula_id = registro.formula.id
    if request.method == 'POST':
        registro.delete()
        messages.success(request, 'Registro eliminado.')
        return redirect('gestionar_registros', formula_id=formula_id)
    
    return render(request, 'sarlaft/eliminar_registro.html', {
        'registro': registro,
        'indicador': registro.formula.indicador,
        'titulo': 'Eliminar Registro'
    })

#? -------------------------------------
@require_POST
def importar_indicadores(request):
    try:
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser un .xlsx')
            return redirect('lista_indicadores')

        df = pd.read_excel(excel_file)

        required_columns = ['nombre', 'formula', 'meta', 'frecuencia_medicion']
        if not all(col in df.columns for col in required_columns):
            messages.error(request, f'El archivo de Excel debe contener las siguientes columnas: {", ".join(required_columns)}')
            return redirect('lista_indicadores')

        for index, row in df.iterrows():
            if any(pd.isna(row[col]) for col in required_columns):
                messages.warning(request, f'Fila {index + 2} omitida por tener valores nulos en campos requeridos.')
                continue

            indicador, created = Indicador.objects.update_or_create(
                nombre=row['nombre'],
                defaults={'orden': index}
            )
            Formula.objects.update_or_create(
                indicador=indicador,
                descripcion=row['formula'],
                defaults={
                    'meta': row['meta'],
                    'frecuencia_medicion': row['frecuencia_medicion'],
                }
            )

        messages.success(request, 'Los indicadores se han importado correctamente.')

    except Exception as e:
        messages.error(request, f'Ocurrió un error al importar el archivo: {e}')

    return redirect('lista_indicadores')

#? -------------------------------------
@require_POST
def importar_registros(request):
    try:
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser un .xlsx')
            return redirect('lista_indicadores')

        df = pd.read_excel(excel_file)

        required_columns = ['nombre_indicador', 'formula_descripcion']
        if not all(col in df.columns for col in required_columns):
            messages.error(request, f'El archivo de Excel debe contener las siguientes columnas: {", ".join(required_columns)}')
            return redirect('lista_indicadores')

        meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                       'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

        for index, row in df.iterrows():
            nombre_indicador = row['nombre_indicador']
            formula_descripcion = row['formula_descripcion']

            try:
                formula = Formula.objects.get(indicador__nombre=nombre_indicador, descripcion=formula_descripcion)
            except Formula.DoesNotExist:
                messages.warning(request, f'Fila {index + 2}: No se encontró la fórmula con descripción "{formula_descripcion}" para el indicador "{nombre_indicador}". Se omitió la fila.')
                continue

            for col_name in df.columns:
                if col_name not in required_columns:
                    try:
                        año = None
                        mes_nombre_final = None

                        if isinstance(col_name, datetime):
                            año = col_name.year
                            mes_numero = col_name.month
                            mes_nombre_final = meses_nombres[mes_numero - 1]
                        else:
                            mes_nombre_str, año_str = str(col_name).split('-')
                            año = int(año_str)
                            
                            temp_mes_nombre = mes_nombre_str.strip().lower()
                            for nombre_capitalizado in meses_nombres:
                                if nombre_capitalizado.lower() == temp_mes_nombre:
                                    mes_nombre_final = nombre_capitalizado
                                    break
                            
                            if not mes_nombre_final:
                                raise ValueError(f"Mes no válido: {mes_nombre_str}")

                        valor = row[col_name]

                        if pd.notna(valor) and año and mes_nombre_final:
                            RegistroIndicador.objects.update_or_create(
                                formula=formula,
                                año=año,
                                mes=mes_nombre_final,
                                defaults={'valor': str(valor)}
                            )
                    except (ValueError, KeyError, IndexError) as e:
                        messages.warning(request, f'Se ignoró la columna "{col_name}" porque no sigue el formato esperado "Mes-Año" o es inválida. Error: {e}')
                        continue
        
        messages.success(request, 'Los registros se han importado o actualizado correctamente.')

    except Exception as e:
        messages.error(request, f'Ocurrió un error al importar el archivo de registros: {e}')

    return redirect('lista_indicadores')

#* -----------------------------------------------
#* FUNCION PARA EXPORTAR EXCEL DE CALIDAD - DATA
#* -----------------------------------------------
def exp_excel_cumplimiento_data(request):
    if request.method == 'POST':
        indicador = request.POST.get('indicador')
        fecha_actual = datetime.now()

        if not indicador:
            return HttpResponse('Indicador es requerido', status=400)

        print("fecha_actual_cumplimiento:", fecha_actual)

        primer_dia_mes_actual = fecha_actual.replace(day=1)
        ultimo_dia_mes_anterior = primer_dia_mes_actual - timedelta(days=1)

        print("Indicador cumplimiento: ", indicador)
        fecha_formateada = ultimo_dia_mes_anterior.strftime('%Y/%m/%d')

        print("Fecha mes anterior cumplimiento: ", fecha_formateada)

        try:
            db = settings.DATABASES['oracle']
            dsn = f"{db['HOST']}:{db['PORT']}/{db['NAME']}"

            with oracledb.connect(user=db['USER'], password=db['PASSWORD'], dsn=dsn) as connection:
                with connection.cursor() as cursor:
                    output_cursor_var = cursor.var(oracledb.CURSOR)
                    cursor.callproc('SP_INDICADORESDATA', [indicador, fecha_formateada, output_cursor_var])
                    output_cursor = output_cursor_var.getvalue()

                    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename="Indicadores_calidad_data.xlsx"'

                    workbook = Workbook()
                    worksheet = workbook.active
                    worksheet.title = "Informe Ind-Calidad data"

                    encabezados = [
                        "MES", "INDICADOR", "VALOR", "META", "DESCRIPCION", "PERIODICIDAD DE EJECUCION", "REQUIERE PLAN DE ACCION", "SEGUIMIENTO PLAN DE ACCION"
                    ]
                    worksheet.append(encabezados)

                    for row in output_cursor:
                        worksheet.append(row)

                    workbook.save(response)

                    return response

        except oracledb.Error as error:
            print('Error de Oracle:', error)
            return HttpResponse('Error al ejecutar el procedimiento', status=500)

    return HttpResponse('MÃ©todo no permitido', status=405)

#? ------------------------------------------------------
#? FUNCIONES PARA DESCARGAR PLANTILLAS DE EXCEL
#? ------------------------------------------------------

def descargar_plantilla_indicadores(request):
    """
    Genera y sirve un archivo Excel (.xlsx) que sirve como plantilla para la importación de indicadores.
    El archivo contiene las cabeceras requeridas para el proceso de importación.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plantilla Indicadores"

    # Definir las cabeceras que espera el proceso de importación
    headers = ['nombre', 'formula', 'meta', 'frecuencia_medicion']
    sheet.append(headers)

    # Crear la respuesta HTTP con el tipo de contenido para un archivo .xlsx
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_indicadores.xlsx"'

    # Guardar el libro de trabajo en la respuesta
    workbook.save(response)

    return response

def descargar_plantilla_registros(request):
    """
    Genera y sirve un archivo Excel (.xlsx) que sirve como plantilla para la importación de registros de indicadores.
    El archivo contiene cabeceras para identificar el indicador y la fórmula, además de columnas de ejemplo para los meses.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plantilla Registros"

    # Definir las cabeceras estáticas
    headers = ['nombre_indicador', 'formula_descripcion']

    # Generar cabeceras de ejemplo para los próximos 3 meses
    today = datetime.now()
    meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    for i in range(3):
        month = today.month + i
        year = today.year
        if month > 12:
            month -= 12
            year += 1
        
        mes_nombre = meses_nombres[month - 1]
        headers.append(f'{mes_nombre}-{year}')

    sheet.append(headers)

    # Crear la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_registros.xlsx"'

    # Guardar el libro de trabajo en la respuesta
    workbook.save(response)

    return response
