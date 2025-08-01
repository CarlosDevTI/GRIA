from datetime import datetime, date, timedelta
import calendar
import logging
import oracledb
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.conf import settings
from django.shortcuts import render
from django.http import HttpResponse

logger = logging.getLogger(__name__)

# ===============================
# FUNCIÓN REUTILIZABLE
# ===============================
def obtener_datos_fondeo():
    """Obtiene los datos desde Oracle usando SP_FONDEO."""
    try:
        fecha_generacion = datetime.now()
        hoy = fecha_generacion.date()

        # Fechas de corte
        _, ultimo_dia_mes = calendar.monthrange(hoy.year, hoy.month)
        fecha_corte_actual = date(hoy.year, hoy.month, ultimo_dia_mes)
        fecha_corte_anterior = hoy.replace(day=1) - timedelta(days=1)

        # Formato para Oracle
        f_actual = fecha_corte_actual.strftime('%Y/%m/%d')
        f_generacion = fecha_generacion.strftime('%Y/%m/%d')
        f_anterior = fecha_corte_anterior.strftime('%Y/%m/%d')

        # Conexión Oracle
        db = settings.DATABASES['oracle']
        dsn = f"{db['HOST']}:{db['PORT']}/{db['NAME']}"

        with oracledb.connect(user=db['USER'], password=db['PASSWORD'], dsn=dsn) as conn:
            with conn.cursor() as cursor:
                ref_cursor_out = cursor.var(oracledb.CURSOR)
                cursor.callproc('SP_FONDEO', [f_actual, f_generacion, f_anterior, ref_cursor_out])
                cur = ref_cursor_out.getvalue()

                if cur:
                    cols = [c[0] for c in cur.description]
                    return [dict(zip(cols, row)) for row in cur]

        return []

    except Exception as e:
        logger.error(f"Error en obtener_datos_fondeo: {e}", exc_info=True)
        return []

# ===============================
# VISTA PRINCIPAL
# ===============================
def seguimiento_fondeo(request):
    datos = []
    if request.method == "POST":
        datos = obtener_datos_fondeo()
        # print(f"Datos obtenidos: {datos}")  # Debugging output

    return render(request, "fondeos/fondeos_dashboard.html", {
        "fondeo_list": datos,
        "fecha_generacion": datetime.now()
    })

# ===============================
# EXPORTAR A EXCEL
# ===============================
def exportar_fondeo_excel(request):
    datos = obtener_datos_fondeo()

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seguimiento Fondeo"

    # Encabezados
    headers = [
        'Cód.', 'Nombre', 'Ahorra Fácil', 'Con Semilla', 'Ahorra Junior',
        'Con Ahorrito', 'Total Ahorros', 'CDAT', 'Contractual',
        'Total CDAT - Contractual', 'Total Captaciones', 'Aportes',
        'Cartera', 'Capt + Apo - Cart', 'Recaudo Alcaldías'
    ]

    # Estilos encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_align = Alignment(horizontal='center')

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # Llenar datos
    money_format = '#,##0'
    for i, row in enumerate(datos, 2):
        ws.cell(i, 1, row.get('CODSUCURSAL'))
        ws.cell(i, 2, row.get('SUCURSAL'))
        ws.cell(i, 3, float(row.get('AHORRAFACIL') or 0)).number_format = money_format
        ws.cell(i, 4, float(row.get('CONSEMILLA') or 0)).number_format = money_format
        ws.cell(i, 5, float(row.get('JUNIOR') or 0)).number_format = money_format
        ws.cell(i, 6, float(row.get('CONAHORRITO') or 0)).number_format = money_format
        ws.cell(i, 7, float(row.get('TOTALAHORROS') or 0)).number_format = money_format
        ws.cell(i, 8, float(row.get('CDAT') or 0)).number_format = money_format
        ws.cell(i, 9, float(row.get('CONTRACTUAL') or 0)).number_format = money_format
        ws.cell(i, 10, float(row.get('TOTCDATCONTRA') or 0)).number_format = money_format
        ws.cell(i, 11, float(row.get('TOTALCAPTACIONES') or 0)).number_format = money_format
        ws.cell(i, 12, float(row.get('APORTES') or 0)).number_format = money_format
        ws.cell(i, 13, float(row.get('CARTERA') or 0)).number_format = money_format
        ws.cell(i, 14, float(row.get('CAPAPOCART') or 0)).number_format = money_format
        ws.cell(i, 15, float(row.get('RECAUDOALCALDIAS') or 0)).number_format = money_format

    # Ajustar anchos
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = length + 2

    # Respuesta
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="fondeo_{datetime.now().date()}.xlsx"'
    wb.save(resp)
    return resp
